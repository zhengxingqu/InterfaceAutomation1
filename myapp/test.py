# -*- coding:utf-8 -*-
import codecs
from itertools import islice
import xlrd
from openpyxl import load_workbook


def read_xlxs1():
    file_url = "/Users/zhengxq/Desktop/case_staff.xlsx"
    wb = xlrd.open_workbook(filename=file_url)
    sheet1 = wb.sheet_by_index(1)
    rows = sheet1.row_values(1)
    cols = sheet1.col_values(3)
    for content in rows:
        print(content)


def rest_xlxs():
    file_url = "/Users/zhengxq/Desktop/case_staff.xlsx"
    wb = load_workbook(filename=file_url)  ##读取路径
    ws = wb.get_sheet_by_name('staff')  ##读取名字为Sheet1的sheet表
    num = 1

    while 1:  # 设定为死循环
        cell = ws.cell(row=num, column=1).value
        if cell:
            num = num + 1
        else:
            num -= 1
            break
    wb = xlrd.open_workbook(filename=file_url)
    sheet1 = wb.sheet_by_index(1)
    for i in range(1, num):
        rows = sheet1.row_values(1)
        print(rows)



if __name__ == "__main__":
    rest_xlxs()
