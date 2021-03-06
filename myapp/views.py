# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from .serializers import UserSerializer, ProjectSerializer, CaseSerializer, \
    ReportSerializer, ReportDetailSerializer, TimingTaskSerializer
from .models import User, Project, Case, Report, ReportDetail, TimingTask
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import generics
from myapp.common import RestonMessage
from rest_framework import status
from rest_framework import filters, viewsets, mixins
import logging
import traceback
import subprocess
import xlrd
from openpyxl import load_workbook
import sys

reload(sys)
sys.setdefaultencoding('utf-8')

logging.basicConfig(filename='/var/log/all.log', level=logging.INFO)


# Create your views here.
# 查询、新增project接口
class InterfaceProject(generics.ListCreateAPIView):
    queryset = Project.objects.filter(isdelete=True)
    serializer_class = ProjectSerializer


class InterfaceCase(generics.ListCreateAPIView):
    queryset = Case.objects.filter(isdelete=True)
    serializer_class = CaseSerializer


class UpdateProject(generics.RetrieveUpdateAPIView):
    queryset = Project.objects.filter(isdelete=True)
    serializer_class = ProjectSerializer


class UpdateCase(generics.RetrieveUpdateAPIView):
    queryset = Case.objects.filter(isdelete=True)
    serializer_class = CaseSerializer


class DeleteProject(generics.RetrieveDestroyAPIView):
    queryset = Project.objects.filter(isdelete=True)
    serializer_class = ProjectSerializer

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        t = self.perform_destroy(instance)
        if t.status_code == 404:
            return Response(RestonMessage.respon_404(),
                            status=status.HTTP_404_NOT_FOUND)
        elif t.status_code == 417:
            return Response(RestonMessage.respon_417(),
                            status=status.HTTP_417_EXPECTATION_FAILED)
        else:
            return Response(RestonMessage.respon_204(),
                            status=status.HTTP_204_NO_CONTENT)

    def perform_destroy(self, instance):

        try:
            exist_project = Case.objects.filter(isdelete=True,
                                                project_name_id=instance.id)
            if exist_project:
                return Response(RestonMessage.respon_417(),
                                status=status.HTTP_417_EXPECTATION_FAILED)
            else:
                Project.objects.filter(id=instance.id).update(isdelete=False)
                logging.INFO(traceback.format_exc())
                return Response(status=status.HTTP_204_NO_CONTENT)
        except Project.DoesNotExist:
            logging.INFO(traceback.format_exc())
            return Response('被删除的对象不存在', status=status.HTTP_404_NOT_FOUND)


class DeleteCase(generics.RetrieveDestroyAPIView):
    queryset = Case.objects.filter(isdelete=True)
    serializer_class = CaseSerializer

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        t = self.perform_destroy(instance)
        if t.status_code == 404:
            return Response(RestonMessage.respon_404(),
                            status=status.HTTP_404_NOT_FOUND)
        elif t.status_code == 417:
            return Response(RestonMessage.respon_417(),
                            status=status.HTTP_417_EXPECTATION_FAILED)
        else:
            return Response(RestonMessage.respon_204(),
                            status=status.HTTP_204_NO_CONTENT)

    def perform_destroy(self, instance):
        try:
            exist_id = '[' + str(instance.id) + ']'
            exist_cases = Case.objects.filter(isdelete=True,
                                              invoking_other_interface=exist_id)
            if exist_cases:
                return Response(RestonMessage.respon_417(),
                                status=status.HTTP_417_EXPECTATION_FAILED)
            else:
                Case.objects.filter(id=instance.id).update(isdelete=False)
                return Response(status=status.HTTP_204_NO_CONTENT)

        except Exception:
            logging.INFO(traceback.format_exc())
            return Response('删除的对象不存在')


class DeleteProjects(APIView):
    queryset = Project.objects.filter(isdelete=True)
    serializer_class = ProjectSerializer

    def put(self, request):
        return_message = {'code': 200, 'message': '成功'}

        data = request.data
        for i in data["ids"]:
            try:
                Project.objects.filter(id=i).update(isdelete=False)
            except Exception:
                logging.INFO(traceback.format_exc())
        return Response(return_message)


class DeleteCases(APIView):
    queryset = Case.objects.filter(isdelete=True)
    serializer_class = CaseSerializer

    def put(self, request):
        return_message = {'code': 200, 'message': '成功'}

        data = request.data
        for i in data["ids"]:
            try:
                Case.objects.filter(id=i).update(isdelete=False)
            except Exception:
                logging.INFO(traceback.format_exc())
        return Response(return_message)


# 搜索固定地址、项目名
class SearchProject(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = Project.objects.filter(isdelete=True)
    serializer_class = ProjectSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('project_name', 'permanent_address')


# 搜索固定地址、项目名
class SearchCase(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = Case.objects.filter(isdelete=True)
    serializer_class = CaseSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('case_name', 'case_result')


class UserList(generics.ListCreateAPIView):
    queryset = User.objects.filter(isdelete=True)
    serializer_class = UserSerializer


class SearchUser(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = User.objects.filter(isdelete=True)
    serializer_class = UserSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('username', 'iphone')


class UpdateUser(generics.RetrieveUpdateAPIView):
    queryset = User.objects.filter(isdelete=True)
    serializer_class = UserSerializer


class DeleteUser(generics.RetrieveDestroyAPIView):
    queryset = User.objects.filter(isdelete=True)
    serializer_class = UserSerializer

    def perform_destroy(self, instance):
        User.objects.filter(id=instance.id).update(isdelete=False)


class DeleteUsers(APIView):
    queryset = Case.objects.filter(isdelete=True)
    serializer_class = UserSerializer

    def put(self, request):
        return_message = {'code': 200, 'message': '成功'}

        data = request.data
        for i in data["ids"]:
            try:
                User.objects.filter(id=i).update(isdelete=False)
            except Exception:
                logging.INFO(traceback.format_exc())
        return Response(return_message)


class ResetPwd(APIView):
    serializer_class = UserSerializer
    queryset = User.objects.filter(isdelete=False)

    def put(self, request, *args, **kwargs):
        data = request.data
        # hash = hashlib.sha1()
        try:
            # hash.update(bytes(data['password']))
            # password = hash.hexdigest()
            t = User.objects.get(id=data['id'])
            t.set_password(data['password'])
            t.save()
        except Exception:
            logging.INFO(traceback.format_exc())
        return Response(data, status=status.HTTP_200_OK)


class CaseReport(generics.ListAPIView):
    serializer_class = ReportSerializer
    queryset = Report.objects.order_by('-test_time')[0:1]


# 根据用例运行时间搜索报告
class SearchReport(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = ReportDetail.objects.all()
    serializer_class = ReportDetailSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('test_time',)


# 根据用例运行时间搜索pass、fail用例数据
class SearchReports(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('test_time',)


# 获取10次最近的报告
class GetReports(generics.ListAPIView):
    queryset = Report.objects.order_by('-test_time')[0:10]
    serializer_class = ReportSerializer


# 获取报告的详细信息
class ReportDetails(generics.ListAPIView):
    result_param = Report.objects.order_by('-test_time')[0:1]
    queryset = ReportDetail.objects.filter(
        test_time=result_param.get().test_time)
    serializer_class = ReportDetailSerializer


class TimingTasks(generics.ListCreateAPIView):
    queryset = TimingTask.objects.filter(isdelete=True)
    serializer_class = TimingTaskSerializer


class SearchTask(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = TimingTask.objects.all()
    serializer_class = TimingTaskSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('task_name',)


class DeleteTask(generics.RetrieveDestroyAPIView):
    queryset = TimingTask.objects.filter(isdelete=True)
    serializer_class = TimingTaskSerializer

    def perform_destroy(self, instance):
        try:
            TimingTask.objects.filter(id=instance.id).update(isdelete=False)
            return Response(status=status.HTTP_204_NO_CONTENT)

        except Exception:
            logging.INFO(traceback.format_exc())
            return Response('删除的对象不存在')


class DeleteTasks(APIView):
    queryset = TimingTask.objects.filter(isdelete=True)
    serializer_class = TimingTaskSerializer

    def put(self, request):
        return_message = {'code': 200, 'message': '成功'}

        data = request.data
        for i in data["ids"]:
            try:
                TimingTask.objects.filter(id=i).update(isdelete=False)
            except Exception:
                logging.INFO(traceback.format_exc())
        return Response(return_message)


class UpdateTaskStatus(generics.RetrieveDestroyAPIView):
    queryset = TimingTask.objects.filter(isdelete=True)
    serializer_class = TimingTaskSerializer

    def perform_destroy(self, instance):
        if TimingTask.objects.get(id=instance.id).is_stop == u'正常':
            try:
                TimingTask.objects.filter(id=instance.id).update(is_stop=u'停用')
                return Response(status=status.HTTP_204_NO_CONTENT)

            except Exception:
                logging.INFO(traceback.format_exc())
                return Response(u'停用的对象不存在')
        else:
            try:
                TimingTask.objects.filter(id=instance.id).update(is_stop=u'正常')
                return Response(status=status.HTTP_204_NO_CONTENT)

            except Exception:
                logging.INFO(traceback.format_exc())
                return Response(u'启用的对象不存在')


class CopyCase(APIView):
    def post(self, request):
        i = 0
        data = request.data
        case_name = Case.objects.get(isdelete=True,
                                     id=data['case_id']).case_name + str(i)
        request_type = Case.objects.get(isdelete=True,
                                        id=data['case_id']).request_type
        request_param = Case.objects.get(isdelete=True,
                                         id=data['case_id']).request_param
        url = Case.objects.get(isdelete=True, id=data['case_id']).url
        project_name = Case.objects.get(isdelete=True,
                                        id=data['case_id']).project_name
        expected_result = Case.objects.get(isdelete=True,
                                           id=data['case_id']).expected_result
        return_result = Case.objects.get(isdelete=True,
                                         id=data['case_id']).return_result
        case_result = Case.objects.get(isdelete=True,
                                       id=data['case_id']).case_result
        invoking_login = Case.objects.get(isdelete=True,
                                          id=data['case_id']).invoking_login
        invoking_other_interface = Case.objects.get(isdelete=True, id=data[
            'case_id']).invoking_other_interface
        isdelete = Case.objects.get(isdelete=True, id=data[
            'case_id']).isdelete
        try:
            logging.info('复制用例信息')
            Case.objects.create(case_name=case_name,
                                request_param=request_param,
                                request_type=request_type, url=url,
                                project_name=project_name,
                                expected_result=expected_result,
                                return_result=return_result,
                                case_result=case_result,
                                invoking_other_interface=invoking_other_interface,
                                invoking_login=invoking_login,
                                isdelete=isdelete)

        except Exception as e:
            return e
        i += 1
        return Response('success')


# 上传文件,自动生成测试用例
class Upload(APIView):

    def post(self, request):
        global data
        data = request.data
        try:
            logging.info("打开文件")
            with open(data['file'].name, 'wb') as f:
                for line in data['file'].readlines():
                    logging.info('创建前端传输的文件')
                    f.write(line)
        except Exception as e:
            return e
        try:
            # 移动所需文件到opt目录
            subprocess.call(["mv", str(data['file'].name), '/opt'])
        except Exception as e:
            return e
        self.read_xlxs()
        return Response('success')

    def read_xlxs(self):
        # 获取excel表的行数
        logging.info('获取excel表的行数')
        file_url = "/opt/" + data['file'].name
        wb = load_workbook(filename=file_url)  ##读取路径
        ws = wb.get_sheet_by_name('staff')  ##读取名字为Sheet1的sheet表
        num = 1

        while 1:  # 设定为死循环
            # 判断每一行的第一列的单元格的值是否存在
            cell = ws.cell(row=num, column=1).value
            if cell:
                num = num + 1
            else:
                num -= 1
                break
        wb = xlrd.open_workbook(filename=file_url)
        sheet1 = wb.sheet_by_index(1)
        # 获取填写行数的字段信息并创建case
        for i in range(1, num):
            rows = sheet1.row_values(1)
            try:
                # 获取对应字段数据
                logging.info('获取excel中接口字段数据')
                case_name = rows[0]
                request_type = rows[1]
                project_name = rows[2]
                request_param = rows[3]
                expected_result = rows[4]
                url = rows[5]
                invoking_login = rows[6]
                invoking_other_interface = rows[7]
                project_name = Project.objects.get(isdelete=True,
                                                   project_name=project_name)

                try:
                    logging.info("创建测试用例")
                    Case.objects.create(case_name=case_name,
                                        request_type=request_type,
                                        project_name=project_name,
                                        request_param=request_param,
                                        expected_result=expected_result,
                                        url=url,
                                        invoking_login=invoking_login,
                                        invoking_other_interface=invoking_other_interface)
                except Exception as e:
                    logging.info(e)
            except Exception as e:
                return e
        return Response("success")


# 上传文件，自动生成项目数据
class UploadProject(APIView):

    def post(self, request):
        global data
        data = request.data
        try:
            #
            with open(data['file'].name, 'wb') as f:
                for line in data['file'].readlines():
                    logging.info('创建前端传输的文件')
                    f.write(line)
        except Exception as e:
            return e
        try:
            subprocess.call(["mv", "-f", str(data['file'].name), '/opt'])
        except Exception as e:
            return e
        self.read_xlxs()
        return Response('success')

    def read_xlxs(self):
        # 获取excel表的行数
        logging.info('获取excel表的行数')
        file_url = "/opt/" + data['file'].name
        # 读取路径
        wb = load_workbook(filename=file_url)
        # 读取名字为case的sheet表
        ws = wb.get_sheet_by_name('case')
        num = 1

        while 1:  # 设定为死循环
            # 判断每一行的第一列的单元格的值是否存在
            cell = ws.cell(row=num, column=1).value
            if cell:
                num = num + 1
            else:
                num -= 1
                break
        wb = xlrd.open_workbook(filename=file_url)
        sheet1 = wb.sheet_by_index(1)
        # 获取填写行数的字段信息并创建case
        for i in range(1, num):
            rows = sheet1.row_values(1)
            try:
                # 获取对应字段数据
                logging.info('获取excel中接口字段数据')
                project_name = rows[0]
                login_way = rows[1]
                permanent_address = rows[2]
                request_header = rows[3]

                try:
                    logging.info("创建项目信息")
                    Project.objects.create(project_name=project_name,
                                           login_way=login_way,
                                           permanent_address=permanent_address,
                                           request_header=request_header)
                except Exception as e:
                    logging.info(e)
                    print(e)
            except Exception as e:
                return e
        return Response("success")
